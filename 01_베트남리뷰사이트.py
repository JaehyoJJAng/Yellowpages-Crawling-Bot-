from bs4 import BeautifulSoup as bs
import requests as rq
import time , re , os ,pyautogui
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook

class ChromeDriver:
    def __init__(self):
        # options 객체 생성
        self.chrome_options = Options()

        # headless Chrome 선언
        self.chrome_options.add_argument('--headless')

        # 브라우저 꺼짐 방지
        self.chrome_options.add_experimental_option('detach', True)

        self.chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.104 Whale/3.13.131.36 Safari/537.36")

        # 불필요한 에러메시지 없애기
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

        self.service = Service(executable_path=ChromeDriverManager().install())
        self.browser = webdriver.Chrome(service=self.service, options=self.chrome_options)
        self.browser.maximize_window()


class Application:
    def __init__(self):
        # url input
        self.URL = self.input_url()

        # ChromeDriver class의 browser 객체 불러오기
        self.browser = ChromeDriver().browser

        # Page Count
        self.page_cnt = self.get_page_count()

        # 데이터 추출 메소드 호출
        # self.get_content()

    # 데이터추출 메소드
    def get_content(self) -> list:
        # 데이터 담을 변수 선언
        save_data = []

        for page in range(1 , self.page_cnt + 1) :
            # 현재 크롤링 페이지 출력
            print(f"{'=' * 60} {page} 페이지 크롤링 시작 {'=' * 60} ")

            # URL에 쿼리 추가
            url = self.URL + f'?page={page}'

            # 페이지 이동
            self.browser.get(url=url)
            self.browser.implicitly_wait(15)

            # html decode 처리
            html = self.browser.page_source.encode('utf-8').decode('utf-8')

            soup = bs(html , 'html.parser')

            # Content 리스트 길이 추출
            contents_len = len(soup.select('div.listing_box'))

            for idx in range(contents_len):
                contents = soup.select('div.listing_box')

                # 회사이름
                title = contents[idx].select_one('h2.company_name > a')
                if title == None or title.text == '':
                    title = '-'
                else:
                    title = str(title.text.strip())

                # 베트남 회사 주소
                address = contents[idx].select_one('p.listing_diachi')
                if address == None or address.text == '':
                    address = '-'
                else:
                    address = address.text.strip()

                # 전화번호
                telephone = contents[idx].select_one('p.listing_tel')
                if telephone == None or telephone.text == '':
                    telephone = '-'
                else:
                    telephone = telephone.text.strip()

                # 이메일 주소
                email = contents[idx].select_one('p.listing_email')
                if email == None or email.text == '':
                    email = '-'
                else:
                    email = email.text.strip()

                # 웹사이트 주소
                website = contents[idx].select('div.listing_website > a')

                if len(website) == 2:
                    website = website[0].attrs['href']
                else:
                    website = '-'

                # 데이터 리스트변수에 저장
                save_data.append([title, address, telephone, email, website])

                # 추출데이터 출력
                print(f"회사이름 : {title}\n주소 : {address}\n전화번호 : {telephone}\n이메일 주소 : {email}\n웹사이트 : {website}\n")

                time.sleep(0.5)

        return save_data

    # 마지막 페이지 숫자 추출 메소드
    def get_page_count(self):
        self.browser.get(url=self.URL)
        self.browser.implicitly_wait(15)

        html = self.browser.page_source.encode('utf-8').decode('utf-8')

        soup = bs(self.browser.page_source,'html.parser')

        try:
            pages = int(soup.select('div#paging > a')[-2].text.strip())
        except :
            pages = 1

        return pages

    # url 입력 메소드
    def input_url(self):
        os.system('cls')
        while True:
            url = input('URL 주소를 입력하세요\nEx ) https://www.yellowpages.vn/cls/77080/phong-kham-nha-khoa.html\n\n:')
            if not url:
                pyautogui.alert("URL 주소가 입력되지 않았습니다!")
                os.system('cls')
                continue

            if len(url) < 10:
                pyautogui.alert(f'URL 형식이 올바르지 않습니다!\n\n입력된 URL\n:{url}')
                os.system('cls')
                continue

            os.system('cls')
            return url


class OpenPyXL(Application):
    def __init__(self):
        super().__init__()
        # Workbook 객체 생성
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(['company name','address','phone number','email address','website'])

        # Application class에서 추출데이터 가져오기
        self.results = self.get_content()

        # 행 초기 값 정의
        self.row = 2

        # 저장경로 지정
        self.savePath = os.path.abspath('yellowpages.vn')

        # 파일이름 짖어
        self.fileName = self.URL.split('/')[-1].replace('.html','') + '.xlsx'

        # 파일저장 메소드 실행
        self.savefile()


    def savefile(self):
        for result in self.results :
            self.ws[f"A{self.row}"] = result[0]
            self.ws[f"B{self.row}"] = result[1]
            self.ws[f"C{self.row}"] = result[2]
            self.ws[f"D{self.row}"] = result[3]
            self.ws[f"E{self.row}"] = result[-1]

            self.row += 1

        if not os.path.exists(self.savePath):
            os.mkdir(self.savePath)

        self.wb.save(os.path.join(self.savePath,self.fileName))
        self.wb.close()

        pyautogui.alert(f'파일 저장 완료!\n\n{self.savePath}')



if __name__ == '__main__' :
    app = OpenPyXL()
